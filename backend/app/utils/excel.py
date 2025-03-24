import os
from typing import List, Dict, Optional, Tuple, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import json
import datetime

from app.schemas.career_import import CareerExcelTemplate, ColumnDefinition, ImportError

# Excel模板列定义
EXCEL_COLUMNS = [
    ColumnDefinition(
        name="title",
        required=True,
        description="职位标题",
        example="Python开发工程师"
    ),
    ColumnDefinition(
        name="company",
        required=True,
        description="公司名称",
        example="XX科技有限公司"
    ),
    ColumnDefinition(
        name="description",
        required=True,
        description="职位描述",
        example="负责公司核心业务系统的开发..."
    ),
    ColumnDefinition(
        name="requirements",
        required=True,
        description="职位要求",
        example="1. 熟练掌握Python编程..."
    ),
    ColumnDefinition(
        name="skills",
        required=False,
        description="所需技能（用逗号分隔）",
        example="Python,Django,FastAPI"
    ),
    ColumnDefinition(
        name="benefits",
        required=False,
        description="职位福利（用逗号分隔）",
        example="五险一金,年终奖,带薪休假"
    ),
    ColumnDefinition(
        name="salary_range",
        required=True,
        description="薪资范围",
        example="15k-25k"
    ),
    ColumnDefinition(
        name="location",
        required=True,
        description="工作地点",
        example="北京"
    ),
    ColumnDefinition(
        name="job_type",
        required=True,
        description="工作类型",
        example="全职"
    ),
    ColumnDefinition(
        name="category_id",
        required=True,
        description="职位分类ID",
        example="1"
    ),
    ColumnDefinition(
        name="experience_required",
        required=True,
        description="所需工作经验",
        example="3-5年"
    ),
    ColumnDefinition(
        name="education_required",
        required=True,
        description="学历要求",
        example="本科"
    ),
]

# 职业Excel模板列定义
CAREER_EXCEL_COLUMNS = [
    ColumnDefinition(
        name="title",
        required=True,
        description="职业名称",
        example="前端开发工程师"
    ),
    ColumnDefinition(
        name="description",
        required=True,
        description="职业描述",
        example="负责网站前端开发和维护..."
    ),
    ColumnDefinition(
        name="required_skills",
        required=True,
        description="所需技能（用逗号分隔）",
        example="HTML, CSS, JavaScript, React"
    ),
    ColumnDefinition(
        name="education_required",
        required=False,
        description="学历要求",
        example="本科"
    ),
    ColumnDefinition(
        name="experience_required",
        required=False,
        description="经验要求",
        example="3-5年"
    ),
    ColumnDefinition(
        name="average_salary",
        required=False,
        description="平均薪资",
        example="15k-25k"
    ),
    ColumnDefinition(
        name="job_outlook",
        required=False,
        description="就业前景",
        example="需求旺盛"
    ),
    ColumnDefinition(
        name="related_majors",
        required=False,
        description="相关专业（用逗号分隔）",
        example="计算机科学,软件工程"
    ),
    ColumnDefinition(
        name="work_activities",
        required=False,
        description="工作内容（用逗号分隔）",
        example="页面开发,性能优化,用户体验改进"
    ),
    ColumnDefinition(
        name="category_name",
        required=True,
        description="职业分类",
        example="软件开发"
    ),
]

def create_template(file_path: str = "job_import_template.xlsx") -> str:
    """创建Excel导入模板
    
    Args:
        file_path: 模板文件保存路径，默认为 job_import_template.xlsx
        
    Returns:
        str: 模板文件保存路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "职位信息"

    # 设置表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # 写入表头
    headers = ["字段名", "是否必填", "描述", "示例值"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    # 写入字段信息
    for row, column in enumerate(EXCEL_COLUMNS, 2):
        ws.cell(row=row, column=1, value=column.name)
        ws.cell(row=row, column=2, value="是" if column.required else "否")
        ws.cell(row=row, column=3, value=column.description)
        ws.cell(row=row, column=4, value=column.example)

    # 调整列宽
    for col in range(1, 5):
        ws.column_dimensions[chr(64 + col)].width = 20

    # 创建数据sheet
    data_sheet = wb.create_sheet("导入数据")
    # 写入表头
    for col, column in enumerate(EXCEL_COLUMNS, 1):
        cell = data_sheet.cell(row=1, column=col)
        cell.value = column.name
        cell.font = header_font
        cell.fill = header_fill
        data_sheet.column_dimensions[chr(64 + col)].width = 20

    wb.save(file_path)
    return file_path

def validate_excel(file_path: str) -> Tuple[List[Dict], List[Dict]]:
    """验证Excel文件并返回有效的数据和错误信息"""
    df = pd.read_excel(file_path, sheet_name="导入数据")
    
    valid_data = []
    errors = []
    required_columns = [col.name for col in EXCEL_COLUMNS if col.required]

    # 检查必填列是否存在
    for col in required_columns:
        if col not in df.columns:
            errors.append({
                "row": 0,
                "column": col,
                "value": "",
                "message": f"缺少必填列: {col}"
            })
    
    if errors:
        return [], errors

    # 验证每一行数据
    for index, row in df.iterrows():
        row_number = index + 2  # Excel行号从2开始（1是表头）
        row_errors = []
        row_data = {}

        # 检查必填字段
        for col in required_columns:
            value = row[col]
            if pd.isna(value) or str(value).strip() == "":
                row_errors.append({
                    "row": row_number,
                    "column": col,
                    "value": "",
                    "message": f"必填字段不能为空"
                })
                continue
            row_data[col] = str(value).strip()

        # 处理可选字段
        for col in df.columns:
            if col not in required_columns:
                value = row[col]
                if not pd.isna(value) and str(value).strip() != "":
                    row_data[col] = str(value).strip()

        # 特殊处理：分隔符字段转换成列表
        for list_field in ["skills", "benefits"]:
            if list_field in row_data:
                row_data[list_field] = [
                    item.strip() for item in row_data[list_field].split(",") if item.strip()
                ]

        if row_errors:
            errors.extend(row_errors)
        else:
            valid_data.append(row_data)

    return valid_data, errors

def validate_career_excel(file_path: str) -> Tuple[List[Dict], List[Dict]]:
    """验证职业Excel数据"""
    valid_data = []
    errors = []
    
    try:
        # 读取Excel文件
        df = pd.read_excel(file_path, sheet_name=0)
        
        # 检查是否有数据
        if df.empty:
            errors.append(ImportError(
                row=0,
                column="",
                value="",
                message="Excel文件没有数据"
            ).dict())
            return valid_data, errors
        
        # 处理每一行数据
        for index, row in df.iterrows():
            row_num = index + 2  # Excel行号（加2是因为1是标题行，index从0开始）
            row_data = {}
            row_errors = []
            
            # 处理每个字段
            for column in CAREER_EXCEL_COLUMNS:
                excel_column_name = column.description  # Excel中使用的是中文列名
                field_name = column.name  # 模型中使用的是英文字段名
                
                # 检查字段是否存在
                if excel_column_name not in df.columns:
                    row_errors.append(ImportError(
                        row=row_num,
                        column=field_name,
                        value="",
                        message=f"缺少必要列: {excel_column_name}"
                    ).dict())
                    continue
                
                # 获取字段值
                value = row[excel_column_name]
                
                # 检查必填字段
                if column.required and (pd.isna(value) or value == ""):
                    row_errors.append(ImportError(
                        row=row_num,
                        column=field_name,
                        value="",
                        message=f"必填字段'{excel_column_name}'不能为空"
                    ).dict())
                    continue
                
                # 如果是空值，设置为None
                if pd.isna(value):
                    row_data[field_name] = None
                else:
                    # 转换为对应的类型
                    if isinstance(value, (int, float)) and not pd.isna(value):
                        if field_name in ["title", "description", "education_required", 
                                        "experience_required", "average_salary", 
                                        "job_outlook", "category_name"]:
                            row_data[field_name] = str(value)
                        else:
                            row_data[field_name] = value
                    else:
                        row_data[field_name] = value
            
            # 如果有错误，记录错误并跳过此行
            if row_errors:
                errors.extend(row_errors)
                continue
            
            # 添加有效数据
            valid_data.append(row_data)
    
    except Exception as e:
        errors.append(ImportError(
            row=0,
            column="",
            value="",
            message=f"解析Excel文件出错: {str(e)}"
        ).dict())
    
    return valid_data, errors 

def process_excel_file(file_path: str, column_definitions: List[ColumnDefinition]) -> Dict[str, Any]:
    """
    处理Excel文件并根据列定义验证和转换数据
    
    Args:
        file_path: Excel文件路径
        column_definitions: 列定义列表
        
    Returns:
        处理结果，包含成功和错误数据
    """
    try:
        # 读取Excel文件
        df = pd.read_excel(file_path)
        
        # 验证必填列是否存在
        required_columns = [col.name for col in column_definitions if col.required]
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return {
                "success": False,
                "error": f"缺少必填列: {', '.join(missing_columns)}",
                "data": None
            }
        
        # 初始化结果
        result = {
            "success": True,
            "total_records": len(df),
            "valid_records": 0,
            "invalid_records": 0,
            "data": [],
            "errors": []
        }
        
        # 处理每一行
        for index, row in df.iterrows():
            row_data = {}
            row_errors = []
            
            # 处理每一列
            for col_def in column_definitions:
                col_name = col_def.name
                
                # 如果列存在于数据中
                if col_name in df.columns:
                    value = row[col_name]
                    
                    # 检查必填项
                    if col_def.required and (pd.isna(value) or value == ""):
                        row_errors.append({
                            "row": index + 2,  # Excel行号从1开始，标题占一行
                            "column": col_name,
                            "value": "",
                            "message": f"{col_name}是必填项"
                        })
                        continue
                    
                    # 处理空值
                    if pd.isna(value):
                        row_data[col_name] = None
                        continue
                    
                    # 根据数据类型进行转换
                    try:
                        if col_def.data_type == "string":
                            row_data[col_name] = str(value).strip()
                        elif col_def.data_type == "integer":
                            row_data[col_name] = int(value)
                        elif col_def.data_type == "float":
                            row_data[col_name] = float(value)
                        elif col_def.data_type == "boolean":
                            row_data[col_name] = bool(value)
                        elif col_def.data_type == "date":
                            if isinstance(value, datetime.datetime):
                                row_data[col_name] = value.date().isoformat()
                            elif isinstance(value, str):
                                # 尝试解析日期字符串
                                try:
                                    dt = datetime.datetime.strptime(value, "%Y-%m-%d")
                                    row_data[col_name] = dt.date().isoformat()
                                except ValueError:
                                    row_errors.append({
                                        "row": index + 2,
                                        "column": col_name,
                                        "value": value,
                                        "message": f"日期格式无效，应为YYYY-MM-DD"
                                    })
                            else:
                                row_errors.append({
                                    "row": index + 2,
                                    "column": col_name,
                                    "value": str(value),
                                    "message": f"无效的日期值"
                                })
                        elif col_def.data_type == "json":
                            if isinstance(value, str):
                                try:
                                    row_data[col_name] = json.loads(value)
                                except json.JSONDecodeError:
                                    row_errors.append({
                                        "row": index + 2,
                                        "column": col_name,
                                        "value": value,
                                        "message": f"JSON格式无效"
                                    })
                            else:
                                row_data[col_name] = value
                        else:
                            # 默认当作字符串处理
                            row_data[col_name] = str(value).strip()
                    except (ValueError, TypeError) as e:
                        row_errors.append({
                            "row": index + 2,
                            "column": col_name,
                            "value": str(value),
                            "message": f"数据类型转换错误: {str(e)}"
                        })
            
            # 添加处理结果
            if row_errors:
                result["invalid_records"] += 1
                result["errors"].extend(row_errors)
            else:
                result["valid_records"] += 1
                result["data"].append(row_data)
        
        return result
    
    except Exception as e:
        return {
            "success": False,
            "error": f"处理Excel文件时出错: {str(e)}",
            "data": None
        }

def generate_template(column_definitions: List[ColumnDefinition], file_path: str) -> bool:
    """
    根据列定义生成Excel导入模板
    
    Args:
        column_definitions: 列定义列表
        file_path: 保存模板的路径
        
    Returns:
        是否成功生成模板
    """
    try:
        # 创建DataFrame
        df = pd.DataFrame(columns=[col.name for col in column_definitions])
        
        # 写入Excel文件
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='导入模板')
            
            # 获取工作表
            workbook = writer.book
            worksheet = writer.sheets['导入模板']
            
            # 添加列描述作为注释
            for col_idx, col_def in enumerate(column_definitions, start=1):
                cell = worksheet.cell(row=1, column=col_idx)
                if col_def.description:
                    from openpyxl.comments import Comment
                    cell.comment = Comment(
                        text=f"{col_def.description}\n{'(必填)' if col_def.required else '(选填)'}",
                        author="系统"
                    )
        
        return True
    
    except Exception as e:
        print(f"生成模板时出错: {str(e)}")
        return False 