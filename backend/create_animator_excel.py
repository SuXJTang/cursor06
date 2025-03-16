import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def create_career_excel(filename="animator_career_fixed2.xlsx"):
    """
    创建一个包含动漫设计师职业数据的Excel文件
    确保与系统模板完全匹配
    """
    # 首先下载系统模板以确保格式一致
    print("正在创建动漫设计师职业数据文件...")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "职业数据"
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 添加表头 - 使用与模板完全相同的表头和顺序
    headers = ["职业名称", "职业描述", "所需技能（用逗号分隔）", "学历要求", "经验要求", 
              "平均薪资", "就业前景", "相关专业（用逗号分隔）", "工作内容（用逗号分隔）", "职业分类"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        
        # 设置列宽
        ws.column_dimensions[chr(64 + col_idx)].width = 20
    
    # 数据行 - 动漫设计师
    row_data = [
        "动漫设计师",  # 职业名称
        "动漫设计师负责创造和发展动画作品中的角色、场景和视觉元素。他们将故事和概念转化为视觉形象，制作分镜脚本，设计角色模型和动画序列，确保整体风格的一致性和艺术品质。",  # 职业描述
        "Photoshop,Illustrator,绘画基础,角色设计,故事板绘制,动画原理,色彩理论",  # 所需技能
        "本科",  # 学历要求
        "1-3年",  # 经验要求
        "8k-15k",  # 平均薪资
        "稳定增长",  # 就业前景
        "动画学,视觉艺术,美术设计,插画设计",  # 相关专业
        "角色设计,场景设计,分镜头设计,动画制作,原画绘制",  # 工作内容
        "艺术创意"  # 职业分类
    ]
    
    # 添加动漫设计师数据
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=2, column=col_idx, value=value)
    
    # 数据行 - 游戏美术设计师
    row_data2 = [
        "游戏美术设计师",  # 职业名称
        "游戏美术设计师负责创建游戏中的视觉元素，包括角色、环境、UI等。他们将概念转化为数字艺术作品，确保游戏的视觉风格符合项目需求，并与开发团队合作实现艺术资源的游戏化。",  # 职业描述
        "3D建模,材质贴图,角色设计,场景设计,Unity/Unreal引擎,Photoshop,Zbrush",  # 所需技能
        "本科",  # 学历要求
        "2-4年",  # 经验要求
        "10k-20k",  # 平均薪资
        "快速增长",  # 就业前景
        "游戏设计,数字媒体艺术,计算机图形学,动画学",  # 相关专业
        "概念设计,3D建模,UV贴图,材质绘制,动画制作,特效设计",  # 工作内容
        "艺术创意"  # 职业分类
    ]
    
    # 添加游戏美术设计师数据
    for col_idx, value in enumerate(row_data2, 1):
        ws.cell(row=3, column=col_idx, value=value)
    
    # 保存文件
    wb.save(filename)
    print(f"已创建职业数据Excel文件: {filename}")
    return filename

if __name__ == "__main__":
    create_career_excel() 