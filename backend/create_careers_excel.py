import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def create_careers_excel(filename="careers_import.xlsx"):
    """
    创建一个包含多个职业数据的Excel文件，用于批量导入测试
    """
    print("正在创建职业数据Excel文件...")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "职业数据"
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 添加表头 - 根据数据库模型定义字段
    headers = ["职业名称", "职业描述", "所需技能（用逗号分隔）", "学历要求", "经验要求", 
              "薪资范围", "发展前景", "职业分类ID"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        
        # 设置列宽
        ws.column_dimensions[chr(64 + col_idx)].width = 20
    
    # 数据行 - 产品经理
    row1_data = [
        "产品经理",  # 职业名称
        "产品经理负责确定产品的愿景、策略和路线图，协调各部门合作将产品从概念转化为可行的商业产品。",  # 职业描述
        "需求分析,用户研究,项目管理,商业分析,沟通协调",  # 所需技能
        "本科",  # 学历要求
        "3-5年",  # 经验要求
        "{'min': 12000, 'max': 30000, 'currency': 'CNY'}",  # 薪资范围
        "市场需求持续增长",  # 发展前景
        "3"  # 职业分类ID
    ]
    
    # 数据行 - SEO专家
    row2_data = [
        "SEO专家",  # 职业名称
        "SEO专家负责优化网站内容和结构，提高网站在搜索引擎中的排名和可见度，增加自然流量。",  # 职业描述
        "关键词研究,内容优化,链接建设,网站分析,SEO工具使用",  # 所需技能
        "大专",  # 学历要求
        "2-4年",  # 经验要求
        "{'min': 8000, 'max': 18000, 'currency': 'CNY'}",  # 薪资范围
        "稳步发展",  # 发展前景
        "2"  # 职业分类ID
    ]
    
    # 添加数据
    for row_idx, data in enumerate([row1_data, row2_data], 2):
        for col_idx, value in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存文件
    wb.save(filename)
    print(f"已创建职业数据Excel文件: {filename}")
    return filename

if __name__ == "__main__":
    create_careers_excel() 