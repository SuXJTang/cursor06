import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = '职业数据'

# 设置样式
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# 列名定义
columns = [
    {'name': 'title', 'description': '职业名称'},
    {'name': 'description', 'description': '职业描述'},
    {'name': 'required_skills', 'description': '所需技能（用逗号分隔）'},
    {'name': 'education_required', 'description': '学历要求'},
    {'name': 'experience_required', 'description': '经验要求'},
    {'name': 'average_salary', 'description': '平均薪资'},
    {'name': 'job_outlook', 'description': '就业前景'},
    {'name': 'related_majors', 'description': '相关专业（用逗号分隔）'},
    {'name': 'work_activities', 'description': '工作内容（用逗号分隔）'},
    {'name': 'category_name', 'description': '职业分类'},
]

# 添加表头
for col_idx, column in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=column['description'])
    cell.font = header_font
    cell.fill = header_fill
    ws.column_dimensions[chr(64 + col_idx)].width = 20

# 添加示例数据
data = [
    {
        '职业名称': '产品经理',
        '职业描述': '负责产品的规划、设计和管理，确保产品满足用户需求并实现商业目标。',
        '所需技能（用逗号分隔）': '需求分析,产品设计,项目管理,沟通协调,数据分析',
        '学历要求': '本科及以上',
        '经验要求': '3-5年',
        '平均薪资': '15k-30k',
        '就业前景': '市场需求稳定，发展前景良好',
        '相关专业（用逗号分隔）': '计算机科学,市场营销,工商管理',
        '工作内容（用逗号分隔）': '需求收集,产品规划,用户研究,功能设计,项目协调',
        '职业分类': '产品设计'
    },
    {
        '职业名称': 'SEO专家',
        '职业描述': '负责网站搜索引擎优化，提高网站在搜索引擎中的排名和可见性。',
        '所需技能（用逗号分隔）': '搜索引擎优化,内容营销,数据分析,HTML/CSS基础,关键词研究',
        '学历要求': '大专及以上',
        '经验要求': '2-4年',
        '平均薪资': '10k-20k',
        '就业前景': '随着数字营销的发展，需求持续增长',
        '相关专业（用逗号分隔）': '市场营销,计算机科学,新媒体',
        '工作内容（用逗号分隔）': '关键词研究,网站分析,内容优化,外链建设,SEO策略制定',
        '职业分类': '互联网营销'
    }
]

# 写入数据
for row_idx, row_data in enumerate(data, 2):
    for col_idx, column in enumerate(columns, 1):
        ws.cell(row=row_idx, column=col_idx, value=row_data[column['description']])

# 保存工作簿
wb.save('correct_careers_import.xlsx')
print('已创建正确的职业数据Excel文件: correct_careers_import.xlsx') 