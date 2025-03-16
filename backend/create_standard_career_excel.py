import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def create_standard_career_excel(filename="standard_careers_import.xlsx"):
    """
    创建一个与系统导入要求完全匹配的标准职业Excel文件
    """
    print("正在创建标准职业数据Excel文件...")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "导入数据"  # 使用"导入数据"作为sheet名称
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 添加表头 - 根据系统导入要求定义字段
    headers = [
        "title", "description", "required_skills", "education_required", 
        "experience_required", "average_salary", "job_outlook", 
        "related_majors", "work_activities", "category_name"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        
        # 设置列宽
        ws.column_dimensions[chr(64 + col_idx)].width = 20
    
    # 数据行 - 产品经理
    row1_data = [
        "产品经理",  # 职业名称
        "产品经理负责确定产品的愿景、策略和路线图，协调各部门合作将产品从概念转化为可行的商业产品。",  # 职业描述
        "需求分析,用户研究,项目管理,商业分析,沟通协调",  # 所需技能
        "本科",  # 学历要求
        "3-5年",  # 经验要求
        "12k-30k",  # 平均薪资
        "市场需求持续增长",  # 就业前景
        "产品设计,市场营销,商业管理",  # 相关专业
        "需求收集,产品规划,用户调研,协调开发",  # 工作活动
        "产品管理"  # 职业分类名称
    ]
    
    # 数据行 - SEO专家
    row2_data = [
        "SEO专家",  # 职业名称
        "SEO专家负责优化网站内容和结构，提高网站在搜索引擎中的排名和可见度，增加自然流量。",  # 职业描述
        "关键词研究,内容优化,链接建设,网站分析,SEO工具使用",  # 所需技能
        "大专",  # 学历要求
        "2-4年",  # 经验要求
        "8k-18k",  # 平均薪资
        "稳步发展",  # 就业前景
        "网络营销,数字媒体,计算机科学",  # 相关专业
        "网站优化,关键词研究,数据分析,内容策略",  # 工作活动
        "互联网营销"  # 职业分类名称
    ]
    
    # 添加数据
    for row_idx, data in enumerate([row1_data, row2_data], 2):
        for col_idx, value in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存文件
    wb.save(filename)
    print(f"已创建标准职业数据Excel文件: {filename}")
    return filename

if __name__ == "__main__":
    create_standard_career_excel() 